#!/usr/bin/env python3
import os
import re
import psycopg2
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook

# -------------------------------------------------------------
# 1. ENV-Variablen laden
# -------------------------------------------------------------
load_dotenv("a.env")

DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PW")
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_SSLMODE = os.getenv("DB_SSLMODE", "require")

# -------------------------------------------------------------
# 2. Verbindung zur DB
# -------------------------------------------------------------
conn = psycopg2.connect(
    dbname=DB_NAME,
    user=DB_USER,
    password=DB_PASSWORD,
    host=DB_HOST,
    port=int(DB_PORT),
    sslmode=DB_SSLMODE
)
cur = conn.cursor()
print("✅ Mit Neon-DB verbunden")

# -------------------------------------------------------------
# 3. Ligen definieren
# -------------------------------------------------------------
season_no = 34
leagues = ['1a', '2a', '2b', '3a', '3b', '3c', '4a', '4b', '4c', '4d', '5a', '5b', '5c']

total_inserted = 0
total_updated = 0

# -------------------------------------------------------------
# 4. Jede Liga-Datei einlesen
# -------------------------------------------------------------
for league in leagues:
    file_name = f"{season_no}th_Backgammon-championships_{league}_output.xlsx"
    if not os.path.exists(file_name):
        print(f"⚠️ Datei nicht gefunden: {file_name}")
        continue

    print(f"📘 Verarbeite {file_name} ...")
    wb = load_workbook(file_name, data_only=True)
    if "Links" not in wb.sheetnames:
        print(f"⚠️ Kein Sheet 'Links' in {file_name}, übersprungen")
        continue
    ws = wb["Links"]

    # ---- Spieler (Zeilen und Spalten) erfassen ----
    players_row = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            players_row.append(str(val).strip())

    players_col = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            players_col.append(str(val).strip())

    # ---- Mapping Namen → player_id aus DB ----
    cur.execute("SELECT player_name, player_id FROM players")
    player_map = {r[0]: r[1] for r in cur.fetchall()}

    # ---- group_id für diese Liga holen ----
    cur.execute("SELECT group_id FROM groups WHERE season_no = %s AND league = %s", (season_no, league))
    res = cur.fetchone()
    if not res:
        print(f"⚠️ Keine group_id für {league} gefunden, übersprungen.")
        continue
    group_id = res[0]

    inserted = 0
    updated = 0

    # ---- Alle Zellen auslesen ----
    for r, player_name in enumerate(players_col, start=2):
        for c, opp_name in enumerate(players_row, start=2):
            if player_name == opp_name:
                continue
            cell = ws.cell(row=r, column=c)
            if not cell.value:
                continue

            # Sichtbarer Text = match_id (nur Zahlen)
            match_text = str(cell.value).strip()
            match_id_match = re.search(r"\d+", match_text)
            if not match_id_match:
                continue
            match_id = int(match_id_match.group(0))

            # Hyperlink = match_link
            match_link = None
            if cell.hyperlink:
                match_link = cell.hyperlink.target

            # DB-IDs abrufen
            player_id = player_map.get(player_name)
            opponent_id = player_map.get(opp_name)
            if not player_id or not opponent_id:
                print(f"⚠️ [{league}] Spieler-ID fehlt: {player_name} vs {opp_name}")
                continue

            try:
                # Prüfen, ob Match schon existiert (diese Richtung)
                cur.execute("""
                    SELECT id FROM matches
                    WHERE group_id = %s AND player_id = %s AND opponent_id = %s
                """, (group_id, player_id, opponent_id))
                existing = cur.fetchone()

                # Prüfen, ob match_id schon irgendwo anders existiert
                cur.execute("SELECT id FROM matches WHERE match_id = %s", (match_id,))
                existing_matchid = cur.fetchone()

                if existing_matchid and (not existing or existing_matchid[0] != existing[0]):
                    print(f"⚠️ [{league}] match_id {match_id} bereits vergeben "
                          f"({player_name} vs {opp_name}) — übersprungen.")
                    continue

                if existing:
                    cur.execute("""
                        UPDATE matches
                        SET match_id = %s, match_link = %s
                        WHERE id = %s
                    """, (match_id, match_link, existing[0]))
                    updated += 1
                else:
                    cur.execute("""
                        INSERT INTO matches (player_id, opponent_id, group_id, match_id, match_link, switched_flag)
                        VALUES (%s, %s, %s, %s, %s, false)
                        ON CONFLICT DO NOTHING
                    """, (player_id, opponent_id, group_id, match_id, match_link))
                    inserted += 1

            except psycopg2.Error as e:
                print(f"⚠️ [{league}] DB-Fehler bei {player_name} vs {opp_name}: {e}")
                conn.rollback()
                continue

    conn.commit()
    print(f"✅ {league}: {inserted} neue / {updated} aktualisierte Matches")
    total_inserted += inserted
    total_updated += updated

# -------------------------------------------------------------
# 5. Abschluss
# -------------------------------------------------------------
print(f"\n🎯 Gesamt abgeschlossen: {total_inserted} neue, {total_updated} aktualisierte Matches")
cur.close()
conn.close()
print("🔚 Verbindung geschlossen.")
