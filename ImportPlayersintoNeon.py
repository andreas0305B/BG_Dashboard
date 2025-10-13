import os
import pandas as pd
import psycopg2
from openpyxl import load_workbook
from dotenv import load_dotenv

# === 1. .env-Datei laden ===
load_dotenv("a.env")

DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PW")
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_SSLMODE = os.getenv("DB_SSLMODE", "require")


# === 2. Excel-Datei laden ===
excel_path = "34th-Backgammon-Championship_Administration.xlsm"

# Workbook laden
wb = load_workbook(excel_path, data_only=True)
ws = wb.active  # aktives Blatt (oder: wb["Sheetname"])

players_data = []

# Nur Spalte J (10) – ab Zeile 2 (da Zeile 1 = Header)
for row in ws.iter_rows(min_row=2, min_col=10, max_col=10, max_row=ws.max_row):
    cell = row[0]  # weil wir nur eine Spalte haben
    if cell.hyperlink:
        link = cell.hyperlink.target      # tatsächliche URL
        name = cell.value                 # sichtbarer Text in der Zelle
        if name and link:
            players_data.append((name, link))

print(f"✅ {len(players_data)} Spieler gefunden.")
print(players_data[:5])  # Zeigt die ersten 5 zum Prüfen


# === 3. Verbindung zur Neon-Datenbank herstellen ===
conn = psycopg2.connect(
    dbname=DB_NAME,
    user=DB_USER,
    password=DB_PASSWORD,
    host=DB_HOST,
    sslmode=DB_SSLMODE
)
cur = conn.cursor()

# === 3b. Sicherstellen, dass keine doppelten Spieler eingefügt werden ===
try:
    cur.execute("""
        ALTER TABLE players
        ADD CONSTRAINT unique_player_link UNIQUE (player_link);
    """)
    conn.commit()
    print("✅ UNIQUE constraint auf 'player_link' erfolgreich hinzugefügt.")
except psycopg2.errors.DuplicateObject:
    conn.rollback()
    print("ℹ️ UNIQUE constraint existiert bereits – übersprungen.")

# === 4. Daten in Tabelle 'players' einfügen ===
for name, link in players_data:
    cur.execute(
        """
        INSERT INTO players (player_name, player_link)
        VALUES (%s, %s)
        ON CONFLICT DO NOTHING;
        """,
        (name, link)
    )

conn.commit()
cur.close()
conn.close()

print("🎯 Upload abgeschlossen!")
