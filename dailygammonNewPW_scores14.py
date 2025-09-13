"""
DailyGammon Score Synchronizer
------------------------------

This script synchronizes tournament match results between an Excel results table
and DailyGammon (DG). It automates the process of filling in missing match IDs,
fetching match results, and updating scores into the correct table cells.

This script processes match results for a specific league and writes them to Excel. 
This script can also run across multiple leagues when used together with the wrapper script!


Usage:
    - Manual mode (default):
        Simply run the script without arguments. 
        Example: python dailygammon.py
        -> Uses default league "4d" hardwired in the script
        -> Keeps Excel workbook open for manual review

    - Command line / wrapper mode:
        Provide the league as the first argument and optionally '--auto' as the second.
        Example: python dailygammon.py 2b --auto
        -> Processes league "2b"
        -> Closes Excel workbook automatically (needed when running multiple leagues in sequence)

This makes it possible to run the script across multiple leagues 
without changing the source code manually.


Core Concepts:
--------------

1. Match ID Handling
   - If a match_id cell is empty, the script searches DG for invitations
     initiated by the "player" and inserts the found ID into Excel.
   - If a match_id cell is filled, it may be either:
        a) Automatically inserted earlier (normal case)
        b) Manually entered by a moderator (manual ID)
   - Manual IDs are detected by comparing player/opponent order between Excel
     and DG: if reversed, the entry is considered manual.

2. Manual Match IDs
   - Stored separately in `matches_by_hand`.
   - They carry a `switched=True` flag, meaning that for DG lookups the roles
     of player and opponent must be swapped to retrieve results.
   - When writing scores back to Excel, the swapped results are re-switched
     so the table remains consistent from the perspective of the Excel player.

3. Caching
   - Each match_id is requested from DG at most once.
   - A simple dict (`html_cache`) maps { match_id -> html } to reduce load.

4. Idempotence
   - Running the script multiple times does not duplicate work.
   - IDs are inserted only if cells are empty; scores are written only if
     the cell does not already contain a final result (e.g., "11").

5. Score Writing
   - For each resolved match, the correct Excel row and columns are located
     via player/opponent name mapping.
   - Exact (case-insensitive) name matches are preferred.
   - If no exact match is found, a heuristic rule is applied:
       * Check whether one name appears as a substring of the other.
   - If the heuristic is inconclusive, the match is skipped for safety.

6. Safety Rules
   - The script never overwrites an existing score of 11.
   - If names cannot be reliably mapped, the match is skipped instead of
     risking a wrong write.

"""


# ============================================================
# Script Purpose:
# This script automatically updates match results for a DailyGammon league season.
# It connects to DailyGammon with your login credentials, collects all match IDs,
# downloads intermediate/final scores, and writes them into the Excel results file.
#
# Workflow in summary:
#   1. Login to DailyGammon with your credentials
#   2. Read the player list from the Excel "Players" sheet
#   3. Detect already known matches from the "Links" sheet
#   4. Find and insert missing match IDs automatically
#   5. Update "Matches" sheet with intermediate 
#   6. For finished matches, set the final winner score to 11
#
# Excel file requirement:
# - Requires Excel file "<season>th_Backgammon-championships_<league>.xlsm"
#   The corresponding Excel file (e.g. "34th_Backgammon-championships_4d.xlsm")
#   must be located in the same folder as this script.
#
# - Excel sheets used:
#       * "Players" → base player list
#       * "Links"   → references to match IDs
#       * "Matches" → current scores
# - Important: Scores are only updated if the match is not yet marked as finished (11).
#
# Before running, configure:
#   - Your User ID and Password (variables: payload["login"], payload["password"])
#   - Current Season number (variable: saison_nummer, e.g. "34")
#   - League (variable: liga, e.g. "4d")
#
# Required Python libraries:
#   requests, beautifulsoup4
#
# If not installed, run:
#   pip install requests beautifulsoup4 openpyxl
#
# ============================================================

import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import streamlit as st
from dotenv import load_dotenv
from openpyxl import Workbook
import os
import pandas as pd
import sys
import pytz
from datetime import datetime


# --- Login Data ---
load_dotenv(dotenv_path="a.env")
login_url = "http://dailygammon.com/bg/login"
DG_LOGIN = os.getenv("DG_LOGIN", "")
DG_PW = os.getenv("DG_PW", "")
payload = {
    "login": DG_LOGIN,
    "password": DG_PW,
    "save": "1"
}

BASE_URL = "http://dailygammon.com/bg/game/{}/0/list"


# -----------------------
# Streamlit Config & Auswahl
# -----------------------
st.set_page_config(
    page_title="Backgammon Championship",
    layout="wide",
    initial_sidebar_state="auto"
)

# CSS für Streamlit Output
st.markdown(
    """
    <style>
    /* Alles nach oben drücken */
    section.main > div {
        padding-top: 0rem !important;
        margin-top: 0rem !important;
    }
    .main .block-container {
        max-width: 100% !important;
        padding-left: 1rem !important; 
        padding-right: 1rem !important;
        padding-top: 0rem !important;
        margin-top: 0rem !important;
    }

    /* Season Selectbox schmaler */
    div[data-baseweb="select"] {
        max-width: 8em !important;
    }

    /* Tabellen auf volle Breite */
    table {
        width: 100% !important;
    }

    /* Pandas Tabellen kompakt */
    table.dataframe th, table.dataframe td {
        padding: 4px 5px !important;
        line-height: 1.4em !important;
        font-size: 14px !important;
    }

    /* Match ID Matrix & Score Matrix */
    table.match-matrix, table.score-matrix {
        border-collapse: collapse;
        width: 100%;
        table-layout: fixed;
    }

    /* Standard-Zellen */
    table.match-matrix th, table.match-matrix td,
    table.score-matrix th, table.score-matrix td {
        border: 1px solid #ddd;
        padding: 4px;
        white-space: nowrap;
        width: 80px;
        text-align: center;  /* Standard: mittig */
    }

    /* Erste Spalte (Spielernamen) linksbündig + sticky */
    table.match-matrix th:first-child,
    table.match-matrix tbody th:first-child,
    table.score-matrix th:first-child,
    table.score-matrix tbody th:first-child {
        text-align: left;
        font-weight: bold;
        position: sticky;
        left: 0;
        z-index: 1;
    }

    /* Farben nach Systemmodus */
    @media (prefers-color-scheme: dark) {
        table.match-matrix th:first-child,
        table.match-matrix tbody th:first-child,
        table.score-matrix th:first-child,
        table.score-matrix tbody th:first-child {
            background-color: #000000;
            color: #ffffff;
        }
    }

    @media (prefers-color-scheme: light) {
        table.match-matrix th:first-child,
        table.match-matrix tbody th:first-child,
        table.score-matrix th:first-child,
        table.score-matrix tbody th:first-child {
            background-color: #f0f0f0;
            color: #000000;
        }
    }

    /* Header oben fixieren */
    table.match-matrix thead th, table.score-matrix thead th {
        position: sticky;
        top: 0;
        z-index: 2;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# -----------------------
# Header
# -----------------------
st.markdown(
    """
    <h1 style='text-align: center; color: #1F3A93; font-size: 3em;'>
        🎲 Backgammon Championship
    </h1>
    """,
    unsafe_allow_html=True
)

st.markdown("---")



# Default Season
col1, col2 = st.columns([0.6, 5])  # <<< Season schmal (1 Teil), Group breit (4 Teile)

with col1:
    season_input = st.selectbox("Season", ["34"], index=0)

with col2:
    sessions = ["1a", "2a", "2b", "3a", "3b", "3c",
                "4a", "4b", "4c", "4d", "5a", "5b", "5c"]
    selection = st.radio("League + Group", sessions, index=1, horizontal=True)

# -----------------------
# Variablen für Script
# -----------------------
# Saison ist direkt aus Auswahl

saison_nummer = season_input

# Liga: entweder aus Wrapper (sys.argv) oder aus Streamlit
if len(sys.argv) > 1:
    liga = sys.argv[1]
else:
    liga = selection

# detect if script was called from wrapper with '--auto'
AUTO_MODE = "--auto" in sys.argv

# Einheitliche Dateinamen
file = f"{saison_nummer}th_Backgammon-championships_{liga}.xlsm"
output_file = f"{saison_nummer}th_Backgammon-championships_{liga}_output.xlsx"
season = f"{saison_nummer}th-season-{liga}"

# Initialisierung, damit sie immer existieren
df_players = None
df_matches = None
df_links = None

print("="*50)
print(f"▶ Script started – collecting links and data for {season}")
print(f"📂 Results saved in Excel file: {file}")
print("="*50)

# -----------------------------------------------------
# --- Read players from Excel ---
# -----------------------------------------------------
# EXCEL FILE ACCESS (OPENPYXL READ-ONLY PHASE):
# - We first use openpyxl to read the "Players" sheet without opening Excel.
# - The file must exist in the same folder and contain a proper "Players" sheet.
# - Player IDs are extracted from hyperlinks in the first column (format: .../bg/user/<id>).
# - If hyperlinks are missing, that player won't get an ID and will be skipped later.
# -----------------------------------------------------


wb_meta = openpyxl.load_workbook(file, data_only=True)
ws_players = wb_meta["Players"]

players = []
player_ids = {}
for row in ws_players.iter_rows(min_row=2, max_col=1, values_only=False):
    cell = row[0]
    if cell.value:
        name = str(cell.value).strip()
        players.append(name)
# - The script assumes each player cell may contain a hyperlink to their DailyGammon user page.
        if cell.hyperlink:
            url = cell.hyperlink.target
            player_id = url.rsplit("/", 1)[-1]
            player_ids[name] = player_id

wb_meta.close()

# -----------------------------------------------------
# --- Create new Excel workbook via openpyxl ---
# -----------------------------------------------------
# OUTPUT FILE (without Macros):
# - We no longer write into the original .xlsm
# - Instead, we create a clean .xlsx with three sheets
# -----------------------------------------------------
# 1st Run Check: Excel einlesen
# -------------------------
# -----------------------------------------------------
# Workbook / Sheets vorbereiten
# -----------------------------------------------------
if os.path.exists(output_file):
    wb_out = openpyxl.load_workbook(output_file)
    ws_links = wb_out["Links"]
    ws_matches = wb_out["Matches"]    
    df_players = pd.read_excel(output_file, sheet_name="Players")
    df_matches = pd.read_excel(output_file, sheet_name="Matches", header=2, index_col=0)
    df_matches.index.name = None
    df_links = pd.read_excel(output_file, sheet_name="Links", header=0, index_col=0)
    df_links.index.name = None
    if "Players" in wb_out.sheetnames:
        ws_players_out = wb_out["Players"]
    else:
        ws_players_out = wb_out.create_sheet("Players")

    # Kontroll-Sheet
    if "Control" in wb_out.sheetnames:
        ws_control = wb_out["Control"]
    else:
        ws_control = wb_out.create_sheet("Control")

else:
    wb_out = Workbook()
    ws_links = wb_out.active
    ws_links.title = "Links"
    ws_matches = wb_out.create_sheet("Matches")
    ws_players_out = wb_out.create_sheet("Players")
    ws_control = wb_out.create_sheet("Control")

    # Matches-Sheet: erste Spalte = Spieler
    for i, player in enumerate(players, start=4):
        cell = ws_matches.cell(row=i, column=1, value=player)
        player_id = player_ids.get(player)
        if player_id:
            cell.hyperlink = f"http://www.dailygammon.com/bg/user/{player_id}"
            cell.style = "Hyperlink"

    # Players-Sheet
    ws_players_out.cell(row=1, column=1, value="Player")
    for i, player in enumerate(players, start=2):
        cell = ws_players_out.cell(row=i, column=1, value=player)
        player_id = player_ids.get(player)
        if player_id:
            cell.hyperlink = f"http://www.dailygammon.com/bg/user/{player_id}"
            cell.style = "Hyperlink"  # Blau + unterstrichen

    # leeres df_links vorbereiten, damit es später immer existiert
    df_links = pd.DataFrame()
# --- Tab 3: Match ID Matrix ---
df_links_clickable = df_links.copy()
# -----------------------------------------------------
# --- Data structures ---
# -----------------------------------------------------
matches = {}
matches_by_hand = {}
match_id_to_excel = {}
html_cache = {}
finished_by_id = {}

# -----------------------------
# Idempotent: Links-Sheet Setup
# -----------------------------
# Schreibe Player in Spalte A (ab Zeile 2), falls nicht schon vorhanden
for i, player in enumerate(players, start=2):
    if ws_links.cell(row=i, column=1).value is None:
        ws_links.cell(row=i, column=1, value=player)

# Schreibe Opponents in Zeile 1 (ab Spalte B), falls nicht schon vorhanden
for j, opponent in enumerate(players, start=2):
    if ws_links.cell(row=1, column=j).value is None:
        ws_links.cell(row=1, column=j, value=opponent)

# -----------------------------------------------------
# --- Login session ---
# -----------------------------------------------------
# -----------------------------------------------------
# Function: login_session
# Purpose:
#   Opens a persistent HTTP session with DailyGammon,
#   logs in with your credentials, and returns the session
#   so all following requests are authenticated.
# -----------------------------------------------------


def login_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": "Mozilla/5.0"})
    resp = s.post(login_url, data=payload, timeout=30)
    resp.raise_for_status()
    return s
session = login_session()


# -----------------------------------------------------
# --- Collect matches per player ---
# -----------------------------------------------------
# -----------------------------------------------------
# Function: get_player_matches
# Purpose:
#   Collects all matches for a specific player in the
#   given season. It scrapes the DailyGammon user page
#   and extracts:
#     - Opponent name
#     - Opponent ID
#     - Match ID
#
# - Filters table rows by the 'season' string to avoid pulling old matches.
# -----------------------------------------------------

def get_player_matches(session: requests.Session, player_id, season):
    url = f"http://www.dailygammon.com/bg/user/{player_id}"
    r = session.get(url)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    player_matches = []
    for row in soup.find_all("tr"):
        text = row.get_text(" ", strip=True)
        if season not in text:
            continue
        opponent_link = row.find("a", href=re.compile(r"/bg/user/\d+"))
        match_link = row.find("a", href=re.compile(r"/bg/game/\d+/0/"))
        if not opponent_link or not match_link:
            continue
        opponent_name = opponent_link.text.strip()
        opponent_id = re.search(r"/bg/user/(\d+)", opponent_link["href"]).group(1)
        match_id = re.search(r"/bg/game/(\d+)/0/", match_link["href"]).group(1)
        player_matches.append((opponent_name, opponent_id, match_id))
    return player_matches

# -----------------------------------------------------
# --- Helper functions: fetch HTML & extract scores ---
# -----------------------------------------------------
# -----------------------------------------------------
# Function: fetch_list_html
# Purpose:
#   Downloads the HTML page for a specific match ID.
#   Returns the HTML content or None if the request failed.
# -----------------------------------------------------

def fetch_list_html(session: requests.Session, match_id: int) -> str | None:
    url = BASE_URL.format(match_id)
    try:
        resp = session.get(url, timeout=30)
        if not resp.ok or "Please Login" in resp.text:
            return None
        return resp.text
    except requests.RequestException:
        return None

# -----------------------------------------------------
# Function: extract_latest_score
# Purpose:
#   Parses the match HTML page and extracts the latest
#   visible score row for the two players.
#   Returns player names + current scores.
#
# PARSE LATEST SCORE FROM MATCH PAGE:
# - Scans table rows from bottom to top (reversed) to find the most recent score line.
# - Assumes the pattern "<Name> : <Score>" is present on both left and right columns.
# -----------------------------------------------------

def extract_latest_score(html: str, players_list: list[str]):
    soup = BeautifulSoup(html, "html.parser")
    for row in reversed(soup.find_all("tr")):
        text = row.get_text(" ", strip=True)
        if not any(p in text for p in players_list):
            continue
        cells = row.find_all("td")
        if len(cells) >= 3:
            left_text = cells[1].get_text(" ", strip=True)
            right_text = cells[2].get_text(" ", strip=True)
            left_match = re.match(r"(.+?)\s*:\s*(\d+)", left_text)
            right_match = re.match(r"(.+?)\s*:\s*(\d+)", right_text)
            if left_match and right_match:
                left_name, left_score = left_match.groups()
                right_name, right_score = right_match.groups()
                return left_name.strip(), right_name.strip(), int(left_score), int(right_score)
    return None

# -----------------------------------------------------
# Function: map_scores_for_excel
# Purpose:
#   Aligns scores from DailyGammon with the correct order
#   in the Excel sheet.
#   Handles switched cases (player order reversed for manual added matches).
#
# NAME/SCORE ALIGNMENT TO EXCEL:
# - The Excel grid expects "excel_player" vs "excel_opponent" in a fixed orientation.
# - 'switched_flag=True' means the match was manually entered with reversed order
#   (excel_player appears on the right on DailyGammon), so we swap scores here.
# - If names match exactly (case-insensitive), we map directly; otherwise we use a
#   small heuristic (substring check) as a fallback. If unsure, return None (skip write).
# -----------------------------------------------------


def map_scores_for_excel(player, opponent, left_name, right_name, left_score, right_score, switched_flag):
    ln = left_name.strip().lower()
    rn = right_name.strip().lower()
    pn = player.strip().lower()
    on = opponent.strip().lower()

    if switched_flag:
        return right_score, left_score

    if ln == pn and rn == on:
        return left_score, right_score
    if ln == on and rn == pn:
        return right_score, left_score

    # Fallback heuristic if names differ slightly
    if pn in ln or pn in rn or on in ln or on in rn:
        if pn in ln:
            return left_score, right_score
        if pn in rn:
            return right_score, left_score
    return None

if df_players is not None and df_matches is not None and df_links is not None:
    # =============================
    # Custom CSS für CI
    # =============================
    st.markdown(
        """
        <style>
        /* Tabs allgemein */
        .stTabs [role="tablist"] {
            background-color: #F8F9F9;
            border-radius: 12px;
            padding: 6px;
            gap: 10px;
        }

        /* Einzelne Tabs */
        .stTabs [role="tab"] {
            background-color: #ffffff;
            color: #333333;
            border-radius: 8px;
            padding: 6px 12px;
            transition: none;
        }

        /* Kein Hover-Effekt */
        .stTabs [role="tab"]:hover {
            background-color: #ffffff !important;
            color: #333333 !important;
        }

        /* Aktiver Tab */
        .stTabs [aria-selected="true"] {
            background-color: #EAEDED !important;
            color: #000000 !important;
            font-weight: 600;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    # -----------------------
    # Sofortiger Excel-Stand Output für alle Tabs
    # -----------------------
    tab1, tab2, tab3 = st.tabs(["League Table", "Score Matrix", "Match ID Matrix"])

    # Platzhalter nur einmal im session_state anlegen
    if "dg_placeholders" not in st.session_state:
        st.session_state.dg_placeholders = {
            "tab1": tab1.empty(),
            "tab2": tab2.empty(),
            "tab3": tab3.empty(),
        }

    placeholder_tab1 = st.session_state.dg_placeholders["tab1"]
    placeholder_tab2 = st.session_state.dg_placeholders["tab2"]
    placeholder_tab3 = st.session_state.dg_placeholders["tab3"]

    
    # --- Tab 1: League Table ---
    players = df_matches.index.tolist()
    intermediate_scores = {}
    for i, player in enumerate(players):
        for j, opponent in enumerate(players):
            if player == opponent:
                continue
            try:
                col_left = 2 * j
                col_right = 2 * j + 1
                left_score = df_matches.iloc[i, col_left]
                right_score = df_matches.iloc[i, col_right]
                if pd.notna(left_score) and pd.notna(right_score):
                    intermediate_scores[(player, opponent)] = (int(left_score), int(right_score))
            except IndexError:
                continue

    # Build League Stats
    num_players = len(players)
    total_matches_per_player = (num_players - 1) * 2
    stats = []
    for player in players:
        finished_matches = 0
        all_plus = all_minus = all_total = 0
        finished_plus = finished_minus = finished_total = 0
        for opponent in players:
            if player == opponent:
                continue
            key_lr = (player, opponent)
            if key_lr in intermediate_scores:
                s_player, s_opponent = intermediate_scores[key_lr]
                all_plus += s_player
                all_minus += s_opponent
                all_total += s_player - s_opponent
                if s_player == 11 or s_opponent == 11:
                    finished_matches += 1
                    finished_plus += s_player
                    finished_minus += s_opponent
                    finished_total += s_player - s_opponent
            key_rl = (opponent, player)
            if key_rl in intermediate_scores:
                s_opp, s_player = intermediate_scores[key_rl]
                all_plus += s_player
                all_minus += s_opp
                all_total += s_player - s_opp
                if s_player == 11 or s_opp == 11:
                    finished_matches += 1
                    finished_plus += s_player
                    finished_minus += s_opp
                    finished_total += s_player - s_opp
        won = sum(1 for (p,o),(sp,so) in intermediate_scores.items() if ((p==player and sp>so) or (o==player and so>sp)) and (sp==11 or so==11))
        lost = sum(1 for (p,o),(sp,so) in intermediate_scores.items() if ((p==player and sp<so) or (o==player and so<sp)) and (sp==11 or so==11))
        pct_won = round((won / (won + lost) * 100)) if (won + lost) > 0 else "---"
        stats.append([player, f"{finished_matches}/{total_matches_per_player}", won, lost, pct_won,
                      all_plus, all_minus, all_total, finished_plus, finished_minus, finished_total])

    df_stats = pd.DataFrame(stats, columns=["Player", "Finished", "Won", "Lost", "% Won",
                                            "All +","All -","All Total","Finished +","Finished -","Finished Total"])

    # MultiIndex-Spalten
    multi_cols = pd.MultiIndex.from_tuples([
        ("", "Player"),
        ("", "Finished"),
        ("", "Won"),
        ("", "Lost"),
        ("", "% Won"),
        ("All matches", "+"),
        ("All matches", "-"),
        ("All matches", "Total"),
        ("Finished matches", "+"),
        ("Finished matches", "-"),
        ("Finished matches", "Total")
    ])
    df_stats.columns = multi_cols

    # Numerische Spalten für sort
    df_stats[("", "Won")] = pd.to_numeric(df_stats[("", "Won")], errors='coerce').fillna(0)
    df_stats[("Finished matches", "Total")] = pd.to_numeric(df_stats[("Finished matches", "Total")], errors='coerce').fillna(0)
    df_stats[("Finished matches", "+")] = pd.to_numeric(df_stats[("Finished matches", "+")], errors='coerce').fillna(0)

    df_stats = df_stats.sort_values(
        by=[("", "Won"), ("Finished matches", "Total"), ("Finished matches", "+")],
        ascending=[False, False, False]
    ).reset_index(drop=True)

    # Tabelle in Platzhalter schreiben
    with tab1:
        df_stats_html = df_stats.to_html(escape=False, index=False)

        last_modified = os.path.getmtime(output_file)
        tz = pytz.timezone("Europe/Berlin")
        last_modified_dt = datetime.fromtimestamp(last_modified, tz)
        formatted_time = last_modified_dt.strftime("%b %d, %Y %H:%M %Z")

        html = df_stats_html + f"<p style='font-size:12px; color:gray;'>Last updated: {formatted_time}</p>"

        placeholder_tab1.markdown(html, unsafe_allow_html=True)

    # --- Tab 2: Score Matrix ---
    matrix_scores = pd.DataFrame("", index=players, columns=players)
    for i, player in enumerate(players):
        for j, opponent in enumerate(players):
            if player == opponent:
                continue
            try:
                col_left = 2 * j
                col_right = 2 * j + 1
                left_score = df_matches.iloc[i, col_left]
                right_score = df_matches.iloc[i, col_right]
                match_id = df_links.iloc[i, j]
                if pd.notna(left_score) and pd.notna(right_score) and pd.notna(match_id):
                    matrix_scores.at[player, opponent] = (
                        f'<a href="http://dailygammon.com/bg/game/{int(match_id)}/0/list#end" target="_blank">'
                        f'{int(left_score)} : {int(right_score)}</a>'
                    )
            except IndexError:
                continue

    # Minimaler Eingriff für linke Spalte als <th> und eigene CSS-Klasse
    html_table = matrix_scores.to_html(escape=False)
    html_table = html_table.replace('<tr><td>', '<tr><th>') \
                        .replace('</td></tr>', '</th></tr>') \
                        .replace('<table border="1" class="dataframe">', '<table class="score-matrix">')

    placeholder_tab2.markdown(html_table, unsafe_allow_html=True)


    # --- Tab 3: Match ID Matrix ---
    df_links_clickable = df_links.copy()

    # Hyperlinks für Match IDs einfügen
    for col in df_links_clickable.columns:
        df_links_clickable[col] = df_links_clickable[col].apply(
            lambda mid: f'<a href="http://dailygammon.com/bg/game/{int(mid)}/0/list#end" target="_blank">{int(mid)}</a>' 
            if pd.notna(mid) else ""
        )

    # DataFrame als HTML-Tabelle mit eigener Klasse rendern
    html_table = df_links_clickable.to_html(escape=False)
    html_table = html_table.replace(
        '<table border="1" class="dataframe">', 
        '<table class="match-matrix">'
    )
    # Ausgabe in Streamlit
    placeholder_tab3.markdown(html_table, unsafe_allow_html=True)
else:
    st.info("1st run: Initializing tables")
    # Platzhalter erzeugen, damit spätere Updates funktionieren
    tab1, tab2, tab3 = st.tabs(["League Table", "Score Matrix", "Match ID Matrix"])
    
    st.session_state.dg_placeholders = {
        "tab1": tab1.empty(),
        "tab2": tab2.empty(),
        "tab3": tab3.empty(),
    }
    placeholder_tab1 = st.session_state.dg_placeholders["tab1"]
    placeholder_tab2 = st.session_state.dg_placeholders["tab2"]
    placeholder_tab3 = st.session_state.dg_placeholders["tab3"]


# Hyperlinks für Match IDs einfügen
#for col in df_links_clickable.columns:
#    df_links_clickable[col] = df_links_clickable[col].apply(
#        lambda mid: f'<a href="http://dailygammon.com/bg/matches/{int(mid)}#end" target="_blank">{int(mid)}</a>' 
#        if pd.notna(mid) else ""
#    )

# Hyperlinks für Match IDs einfügen (ohne Diagonale: Player vs. sich selbst)
for row_idx, row_name in enumerate(df_links_clickable.index):
    for col_idx, col_name in enumerate(df_links_clickable.columns):
        if row_name == col_name:  
            # Diagonale -> leer
            df_links_clickable.iat[row_idx, col_idx] = ""
        else:
            mid = df_links_clickable.iat[row_idx, col_idx]
            if pd.notna(mid) and str(mid).strip().isdigit():
                mid_int = int(mid)
                df_links_clickable.iat[row_idx, col_idx] = (
                    f'<a href="http://dailygammon.com/bg/game/{int(mid)}/0/list#end" target="_blank">{mid_int}</a>'
                )
            else:
                df_links_clickable.iat[row_idx, col_idx] = ""

# DataFrame als HTML-Tabelle mit eigener Klasse rendern
html_table = df_links_clickable.to_html(escape=False)
html_table = html_table.replace(
    '<table border="1" class="dataframe">', 
    '<table class="match-matrix">'
)

# Ausgabe in Streamlit
placeholder_tab3.markdown(html_table, unsafe_allow_html=True)


# Extract players/columns from "Links"
# "LINKS" SHEET LAYOUT ASSUMPTION:
# - Column A (from row 2 down) lists row player names.
# - Row 1 (from column B rightwards) lists opponent names (as columns).
# - Cells at (row_player, col_opponent) hold the match ID (and hyperlink).

row_players_links = []
r = 2
while True:
    v = ws_links.cell(row=r, column=1).value  # Spalte A
    if not v:
        break
    row_players_links.append(str(v).strip())
    r += 1

col_opponents_links = []
c = 2
while True:
    v = ws_links.cell(row=1, column=c).value  # Zeile 1, ab Spalte B
    if not v:
        break
    col_opponents_links.append(str(v).strip())
    c += 1

col_index_links = {name: 2 + i for i, name in enumerate(col_opponents_links)}
# -----------------------------------------------------
# Step 1 (light): Build match_id_to_excel from Excel (no web fetch)
# -----------------------------------------------------
match_id_to_excel = {}
for i, player_name in enumerate(row_players_links, start=2):
    for opp in col_opponents_links:
        if player_name == opp:
            continue
        c = col_index_links.get(opp)
        val = ws_links.cell(row=i, column=c).value
        if not val:
            continue
        try:
            mid = int(val)
        except Exception:
            try:
                mid = int(str(val).strip())
            except Exception:
                continue
        match_id_to_excel[mid] = (player_name, opp, False)

# -----------------------------------------------------
# Step 1: Check existing links with HTML fetch (optimized)
# -----------------------------------------------------
# 1. Sammle alle Match-IDs, die noch nicht im Cache sind
to_fetch_ids = []
for i, player_name in enumerate(row_players_links, start=2):
    for opp in col_opponents_links:
        if player_name == opp:
            continue
        c = col_index_links.get(opp)
        val = ws_links.cell(row=i, column=c).value
        if not val:
            continue
        try:
            match_id = int(val)
        except Exception:
            match_id = int(str(val).strip())
        if match_id not in html_cache and match_id not in finished_by_id:
            to_fetch_ids.append(match_id)

# 2. Batch fetch
for match_id in to_fetch_ids:
    html_cache[match_id] = fetch_list_html(session, match_id)

# 3. Prüfe alle Matches
for i, player_name in enumerate(row_players_links, start=2):
    for opp in col_opponents_links:
        if player_name == opp:
            continue
        c = col_index_links.get(opp)
        val = ws_links.cell(row=i, column=c).value
        if not val:
            continue
        try:
            match_id = int(val)
        except Exception:
            match_id = int(str(val).strip())

        # Fertige Matches überspringen
        if match_id in finished_by_id:
            matches[(player_name, opp)] = match_id
            match_id_to_excel[match_id] = (player_name, opp, False)
            continue

        html = html_cache.get(match_id)
        if not html:
            matches[(player_name, opp)] = match_id
            match_id_to_excel[match_id] = (player_name, opp, False)
            continue

        score_info = extract_latest_score(html, [player_name, opp])
        if not score_info:
            matches[(player_name, opp)] = match_id
            match_id_to_excel[match_id] = (player_name, opp, False)
            continue

        left_name, right_name, status, _ = score_info
        ln, rn = left_name.lower(), right_name.lower()
        pn, on = player_name.lower(), opp.lower()

        if status == "finished":
            finished_by_id[match_id] = True

        if ln == pn and rn == on:
            matches[(player_name, opp)] = match_id
            match_id_to_excel[match_id] = (player_name, opp, False)
        elif ln == on and rn == pn:
            matches_by_hand[(player_name, opp)] = (match_id, True)
            match_id_to_excel[match_id] = (player_name, opp, True)
            print(f"Found manual inserted match detected: {player_name} vs {opp} with match ID {match_id}.")
        else:
            matches[(player_name, opp)] = match_id
            match_id_to_excel[match_id] = (player_name, opp, False)
            print(f"⚠️ Unclear order for match ID {match_id}: DG shows '{left_name}' vs '{right_name}'")

# -----------------------------------------------------
# Step 2: Fill missing match IDs (with skip-check)
# -----------------------------------------------------
flag_cell = ws_control["A1"].value
if flag_cell == "MATCH_IDS_FILLED":
    skip_id_fetch = True
    print("✅ All match IDs are marked as filled — skipping DailyGammon fetch.")
else:
    skip_id_fetch = False

if not skip_id_fetch:
    for player in players:
        pid = player_ids.get(player)
        if not pid:
            continue
        missing = [opp for opp in players if opp != player and (player, opp) not in matches and (player, opp) not in matches_by_hand]
        if not missing:
            continue
        player_matches = get_player_matches(session, pid, season=season)
        for opponent_name, opponent_id, match_id in player_matches:
            key = (player, opponent_name)
            if key in matches or key in matches_by_hand:
                continue
            mid_int = int(match_id)
            switched_flag = False
            if mid_int in match_id_to_excel:
                _, _, switched_flag = match_id_to_excel[mid_int]
            matches[key] = mid_int
            match_id_to_excel[mid_int] = (player, opponent_name, switched_flag)
            try:
                row_idx = row_players_links.index(player) + 2
            except ValueError:
                continue
            c = col_index_links.get(opponent_name)
            if not c or opponent_name == player:
                continue
            cell = ws_links.cell(row=row_idx, column=c)
            if not cell.value:
                cell.value = str(match_id)
                cell.hyperlink = f"http://www.dailygammon.com/bg/game/{match_id}/0/list#end"
                cell.style = "Hyperlink"
                print(f"Detected missing match between {player} and {opponent_name} — match ID={match_id} has been auto-added to the table")

    print("✅ Match IDs updated (auto + manual detection)")
else:
    print("ℹ️ Skipping Step 2: All match IDs already present in Excel")

ws_control["A1"].value = "MATCH_IDS_FILLED"

# -----------------------------------------------------
# Step 3: Collect finished matches
# Purpose:
#   For every player, fetch their export page.
#   If a match is marked as finished, extract the winner.
#   Results are stored in a dictionary for later processing.
#
# FINISHED MATCH DETECTION:
# - We open each player's page and follow "export" links for matches of this season.
# - The winner is inferred from a simple textual rule (position of "Wins" on the line).
# - 'finished_by_id' maps match_id -> winner_name for later use in Phase 2.
# -----------------------------------------------------

for player in players:
    pid = player_ids.get(player)
    if not pid:
        continue
    url = f"http://www.dailygammon.com/bg/user/{pid}"
    try:
        r = session.get(url, timeout=30)
        r.raise_for_status()
    except requests.RequestException:
        continue
    soup = BeautifulSoup(r.text, "html.parser")
    for row in soup.find_all("tr"):
        text = row.get_text(" ", strip=True)
        if season not in text:
            continue
        export_link = row.find("a", href=re.compile(r"/bg/export/\d+"))
        match_link = row.find("a", href=re.compile(r"/bg/game/\d+/0/"))
        opponent_link = row.find("a", href=re.compile(r"/bg/user/\d+"))
        if not export_link or not match_link or not opponent_link:
            continue
        try:
            match_id = int(re.search(r"/bg/game/(\d+)/0/", match_link["href"]).group(1))
        except Exception:
            continue
        opponent_name = opponent_link.text.strip()
        export_url = f"http://www.dailygammon.com/bg/export/{match_id}"
        try:
            resp_export = session.get(export_url, timeout=30)
            text_lines = resp_export.text.splitlines()
        except requests.RequestException:
            continue
        winner = None

# - 'mid_threshold 24' is a rough character-position cutoff to decide whether the "Wins"
#   belongs to the left or right player on the export line.

        mid_threshold = 24
        for line in text_lines:
            if "and the match" in line and "Wins" in line:
                pos = line.find("Wins")
                winner = player if pos < mid_threshold else opponent_name
                break
        if winner:
            finished_by_id[match_id] = winner

# -----------------------------------------------------
# Phase 1: Write intermediate scores
# Purpose:
#   For each match, download the latest score and update
#   the "Matches" sheet in Excel.
#   IMPORTANT: If a score of 11 is already present,
#   the match is considered finished and will not be overwritten.
# -----------------------------------------------------

print("🔎 Phase 1: Writing intermediate scores for matches...")
players_in_matches = []
row_counter = 4
while True:
    nm = ws_matches.cell(row=row_counter,column=1).value
    if not nm:
        break
    players_in_matches.append(str(nm).strip())
    row_counter += 1
col_start = 2

# EXCEL WRITE HELPER (INTERMEDIATE SCORES):
# - Translates (excel_player, excel_opponent) to row/column indices in "Matches".
# - For each opponent, we reserve two columns: left=excel_player's score, right=excel_opponent's score.
# - Safety: if either cell already equals 11, we skip to preserve final results.
# - Scores are already correctly oriented by 'map_scores_for_excel'; no swapping here.

def write_score_to_excel(excel_player, excel_opponent, player_score, opponent_score, switched_flag):
    try:
        r_idx = players_in_matches.index(excel_player) + 4
        c_base = players_in_matches.index(excel_opponent)
    except ValueError:
        print(f"⚠️ Player not found in Excel sheet: {excel_player} vs {excel_opponent}")
        return False
    c_left = col_start + c_base * 2
    c_right = c_left + 1

    # Do not overwrite already finished (11) scores!
    # - Once a match is finished (11), intermediate updates must never overwrite that cell.

    left_cell_val = ws_matches.cell(row=r_idx, column=c_left).value
    right_cell_val = ws_matches.cell(row=r_idx, column=c_right).value
    if left_cell_val == 11 or right_cell_val == 11:
        return False

    ws_matches.cell(row=r_idx, column=c_left, value=player_score)
    ws_matches.cell(row=r_idx, column=c_right, value=opponent_score)
    return True

# - Pull HTML from cache if available; otherwise fetch fresh.
for match_id, (excel_player, excel_opponent, switched_flag) in list(match_id_to_excel.items()):
    # 🔍 1. Check in Excel, ob das Match schon beendet ist
    try:
        row_idx = row_players_links.index(excel_player) + 2
        col_idx = col_index_links.get(excel_opponent)

        left_score = ws_matches.cell(row=row_idx, column=col_idx*2 - 1).value
        right_score = ws_matches.cell(row=row_idx, column=col_idx*2).value

        if left_score == 11 or right_score == 11:
            # ✅ Match schon beendet -> kein Fetch mehr nötig
            matches[(excel_player, excel_opponent)] = match_id
            continue
    except Exception:
        # falls Spieler oder Spalte nicht gefunden wird -> normal weitermachen
        pass

    # 🔍 2. Nur wenn das Match noch offen ist -> HTML ziehen
    html = html_cache.get(match_id)
    if not html:
        html = fetch_list_html(session, match_id)
        html_cache[match_id] = html
    if not html:
        continue

    result = extract_latest_score(html, players_in_matches)
    if not result:
        continue

    left_name, right_name, left_score, right_score = result
    # ... dein bisheriger Code für switched / Zuordnung ...

    # Map scores based on player names

    mapped = map_scores_for_excel(excel_player, excel_opponent, left_name, right_name, left_score, right_score, switched_flag)
    if mapped is None:
        continue
    excel_player_score, excel_opponent_score = mapped
    write_score_to_excel(excel_player, excel_opponent, excel_player_score, excel_opponent_score, switched_flag)

print("✅ Phase 1: completed")

# -----------------------------------------------------
# Phase 2: Final results - Set winners to 11 points
# Purpose:
#   For matches identified as finished, write the final
#   winner score (11 points) into the correct player cell
#   in the "Matches" sheet.
# -----------------------------------------------------

print("🔎 Phase 2: Final results (set winner = 11) ...")
for match_id, winner_name in finished_by_id.items():
    info = match_id_to_excel.get(match_id)
    if not info:
        continue
    excel_player, excel_opponent, switched_flag = info
    try:
        r_idx = players_in_matches.index(excel_player) + 4
        c_base = players_in_matches.index(excel_opponent)
    except ValueError:
        continue
    c_left = col_start + c_base * 2
    c_right = c_left + 1

    # Write 11 to the correct winner cell

    winner_lower = winner_name.strip().lower()
    if switched_flag:
        if winner_lower == excel_player.lower():
            ws_matches.cell(row=r_idx, column=c_right, value=11)
        elif winner_lower == excel_opponent.lower():
            ws_matches.cell(row=r_idx, column=c_left, value=11)
    else:
        if winner_lower == excel_player.lower():
            ws_matches.cell(row=r_idx, column=c_left, value=11)
        elif winner_lower == excel_opponent.lower():
            ws_matches.cell(row=r_idx, column=c_right, value=11)

wb_out.save(output_file)
# Close workbook automatically only if called from wrapper
if AUTO_MODE:
    wb_out.close()
print("🏁 Script finished successfully")
print("="*50)


# -----------------------
# Load Excel Sheets
# -----------------------
df_players = pd.read_excel(output_file, sheet_name="Players")
df_matches = pd.read_excel(output_file, sheet_name="Matches", header=2, index_col=0)  # Index = erste Spalte
df_matches.index.name = None  # Entfernt "Unnamed: 0"
df_links = pd.read_excel(output_file, sheet_name="Links", header=0, index_col=0)      # Index = erste Spalte
df_links.index.name = None  # Entfernt "Unnamed: 0"

# -----------------------
# CSS für Tabs
# -----------------------
st.markdown("""
<style>
div[role="tab"] { font-size: 14px !important; padding: 4px 8px !important; }
div[role="tablist"] { margin-bottom: 10px !important; }
</style>
""", unsafe_allow_html=True)

# -----------------------
# Tabs aktualisieren
# -----------------------

with tab2:
    # Spielerliste aus Index der Matches-Tabelle
    players = df_matches.index.tolist()
    opponents = players  # gleiche Reihenfolge wie Index

    # Neue DataFrame für sauber formatiertes Ergebnis
    matrix_scores = pd.DataFrame("", index=players, columns=opponents)

    for i, player in enumerate(players):
        for j, opponent in enumerate(opponents):
            if player == opponent:
                matrix_scores.at[player, opponent] = ""  # keine Spiele gegen sich selbst
                continue

            # Spalten im Matches-DataFrame: links = player, rechts = opponent
            col_left = 2 * j
            col_right = 2 * j + 1

            try:
                left_score = df_matches.iloc[i, col_left]
                right_score = df_matches.iloc[i, col_right]

                # Match-ID aus df_links: gleiche Zeile/Spalte wie Spieler vs Gegner
                match_id = df_links.iloc[i, j]  # Indexspalte wird korrekt berücksichtigt

                if pd.notna(left_score) and pd.notna(right_score) and pd.notna(match_id):
                    score_text = f"{int(left_score)} : {int(right_score)}"
                    # Hyperlink für das Match einfügen
                    matrix_scores.at[player, opponent] = (
                        f'<a href="http://dailygammon.com/bg/game/{int(match_id)}/0/list#end" '
                        f'target="_blank">{score_text}</a>'
                    )
            except IndexError:
                continue

    # Streamlit: HTML-Output, klickbare Scores
    html_table = matrix_scores.to_html(escape=False)
    # linke Spalte als <th>
    html_table = html_table.replace('<tr><td>', '<tr><th>').replace('</td></tr>', '</th></tr>')
    # eigene Klasse für CSS
    html_table = html_table.replace('<table border="1" class="dataframe">', '<table class="score-matrix">')
    placeholder_tab2.markdown(html_table, unsafe_allow_html=True)

with tab3:
    df_links_clickable = df_links.copy()

    # Hyperlinks erzeugen, alle Spalten
    for col in df_links_clickable.columns:
        df_links_clickable[col] = df_links_clickable[col].apply(
            lambda mid: f'<a href="http://dailygammon.com/bg/game/{int(mid)}/0/list#end" target="_blank">{int(mid)}</a>'
            if pd.notna(mid) else ""
        )

    # Index bleibt als Spielername, keine Unnamed: 0 mehr
    html_table = df_links_clickable.to_html(escape=False)
    html_table = html_table.replace('<table border="1" class="dataframe">', '<table class="match-matrix">')
    placeholder_tab3.markdown(html_table, unsafe_allow_html=True)

players = df_matches.index.tolist()
intermediate_scores = {}

for i, player in enumerate(players):
    for j, opponent in enumerate(players):
        if player == opponent:
            continue
        try:
            col_left = 2 * j
            col_right = 2 * j + 1
            left_score = df_matches.iloc[i, col_left]
            right_score = df_matches.iloc[i, col_right]
            if pd.notna(left_score) and pd.notna(right_score):
                intermediate_scores[(player, opponent)] = (int(left_score), int(right_score))
        except IndexError:
            continue

# -----------------------
# Build League Table / Stats for Tab 1
# -----------------------
num_players = len(players)
total_matches_per_player = (num_players - 1) * 2  # home + away
stats = []

for player in players:
    finished_matches = 0
    all_plus = all_minus = all_total = 0
    finished_plus = finished_minus = finished_total = 0

    for opponent in players:
        if player == opponent:
            continue

        # Player as row
        key_lr = (player, opponent)
        if key_lr in intermediate_scores:
            s_player, s_opponent = intermediate_scores[key_lr]
            all_plus += s_player
            all_minus += s_opponent
            all_total += s_player - s_opponent
            if s_player == 11 or s_opponent == 11:
                finished_matches += 1
                finished_plus += s_player
                finished_minus += s_opponent
                finished_total += s_player - s_opponent

        # Player as column
        key_rl = (opponent, player)
        if key_rl in intermediate_scores:
            s_opp, s_player = intermediate_scores[key_rl]
            all_plus += s_player
            all_minus += s_opp
            all_total += s_player - s_opp
            if s_player == 11 or s_opp == 11:
                finished_matches += 1
                finished_plus += s_player
                finished_minus += s_opp
                finished_total += s_player - s_opp

    # Gewonnene/Verlorene Matches
    won = sum(1 for (p,o),(sp,so) in intermediate_scores.items() if ((p==player and sp>so) or (o==player and so>sp)) and (sp==11 or so==11))
    lost = sum(1 for (p,o),(sp,so) in intermediate_scores.items() if ((p==player and sp<so) or (o==player and so<sp)) and (sp==11 or so==11))
    pct_won = round((won / (won + lost) * 100)) if (won + lost) > 0 else "---"

    stats.append([
        player,
        f"{finished_matches} / {total_matches_per_player}",
        won,
        lost,
        pct_won,
        all_plus,
        all_minus,
        all_total,
        finished_plus,
        finished_minus,
        finished_total
    ])

# --- Build DataFrame ---
df_stats = pd.DataFrame(
    stats,
    columns=[
        "Player", "Finished", "Won", "Lost", "% Won",
        "All +", "All -", "All Total",
        "Finished +", "Finished -", "Finished Total"
    ]
)

multi_cols = pd.MultiIndex.from_tuples([
    ("", "Player"),
    ("", "Finished"),
    ("", "Won"),
    ("", "Lost"),
    ("", "% Won"),
    ("All matches", "+"),
    ("All matches", "-"),
    ("All matches", "Total"),
    ("Finished matches", "+"),
    ("Finished matches", "-"),
    ("Finished matches", "Total")
])
df_stats.columns = multi_cols

df_stats = df_stats.sort_values(
    by=[("", "Won"), ("Finished matches", "Total"), ("Finished matches", "+")],
    ascending=[False, False, False]
).reset_index(drop=True)

# --- Render League Table in Streamlit ---

with tab1:
    # Tabelle als HTML
    df_stats_html = df_stats.to_html(escape=False, index=False)

    # Excel Last Modified Timestamp
    last_modified = os.path.getmtime(output_file)
    tz = pytz.timezone("Europe/Berlin")
    last_modified_dt = datetime.fromtimestamp(last_modified, tz)
    formatted_time = last_modified_dt.strftime("%b %d, %Y %H:%M %Z")

    # Tabelle + Timestamp kombiniert in denselben Platzhalter schreiben
    html = df_stats_html + f"<p style='font-size:12px; color:gray;'>Last updated: {formatted_time}</p>"
    placeholder_tab1.markdown(html, unsafe_allow_html=True)
