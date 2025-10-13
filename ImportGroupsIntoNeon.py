#!/usr/bin/env python3
import os
from openpyxl import load_workbook
from dotenv import load_dotenv
import psycopg2
import sys

# -------- 1) .env laden --------
load_dotenv("a.env")

DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PW")
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_SSLMODE = os.getenv("DB_SSLMODE", "require")

if not all([DB_NAME, DB_USER, DB_PASSWORD, DB_HOST]):
    print("❌ Fehlende DB-Umgebungsvariablen. Bitte a.env prüfen.")
    sys.exit(1)

# -------- 2) Excel laden und Sheet Members --------
excel_path = "34th-Backgammon-Championship_Administration.xlsm"
wb = load_workbook(excel_path, data_only=True)

sheet_name = "Members"
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    ws = wb.active
    print(f"⚠️ Arbeitsblatt '{sheet_name}' nicht gefunden — verwende aktives Blatt: '{ws.title}'")

# -------- 3) Season aus C2 --------
season_cell = ws["C2"].value
try:
    season_no = int(season_cell)
except (TypeError, ValueError):
    print(f"❌ Ungültige Season-Nummer in C2: {season_cell!r}")
    sys.exit(1)

print(f"🏁 Season-Nummer erkannt: {season_no}")

# -------- 4) Gruppen aus Spalte D ab Zeile 2 einlesen, Duplikate vermeiden --------
import re

pattern = re.compile(r"^\d+[a-zA-Z]$")  # z.B. 1a, 3b, 10c

groups_set = set()

for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, max_row=ws.max_row):
    cell = row[0]
    if cell.value:
        val = str(cell.value).strip()
        if pattern.match(val):
            groups_set.add(val)

groups = sorted(groups_set)
print(f"✅ {len(groups)} eindeutige Gruppen gefunden: {groups}")

if not groups:
    print("⚠️ Keine Gruppen gefunden. Abbruch.")
    sys.exit(0)

# -------- 5) Verbindung zur Neon-DB --------
try:
    conn = psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=int(DB_PORT),
        sslmode=DB_SSLMODE
    )
except Exception as e:
    print("❌ Verbindung zur DB fehlgeschlagen:", e)
    sys.exit(1)

cur = conn.cursor()

# -------- 6) Gruppen einfügen (ON CONFLICT DO NOTHING) --------
inserted = 0
for group in groups:
    cur.execute(
        """
        INSERT INTO groups (season_no, league)
        VALUES (%s, %s)
        ON CONFLICT DO NOTHING;
        """,
        (season_no, group)
    )
    if cur.rowcount == 1:
        inserted += 1

conn.commit()

print(f"🎯 Import abgeschlossen: {inserted} neue Gruppen eingefügt, {len(groups)-inserted} übersprungen.")

# Ab Zeile 2 einlesen
mappings = set()  # (player_name, season, group_name)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    player_cell = row[9]  # Spalte J
    group_cell = row[3]   # Spalte D
    season_cell = row[2]  # Spalte C

    if not player_cell.value or not group_cell.value or not season_cell.value:
        continue

    player_name = str(player_cell.value).strip()
    group_name = str(group_cell.value).strip()
    season_no = int(season_cell.value)

    if pattern.match(group_name):
        mappings.add((player_name, season_no, group_name))

print(f"✅ {len(mappings)} gültige Player-Group-Zuordnungen gefunden")

# --- IDs aus DB holen und player_groups befüllen ---
inserted = 0
for player_name, season_no, group_name in mappings:
    # player_id
    cur.execute("SELECT player_id FROM players WHERE player_name = %s", (player_name,))
    res = cur.fetchone()
    if not res:
        print(f"⚠️ Player '{player_name}' nicht gefunden, übersprungen")
        continue
    player_id = res[0]

    # group_id
    cur.execute("SELECT group_id FROM groups WHERE season_no = %s AND league = %s", (season_no, group_name))
    res = cur.fetchone()
    if not res:
        print(f"⚠️ Gruppe '{group_name}' in Season {season_no} nicht gefunden, übersprungen")
        continue
    group_id = res[0]

    # Einfügen
    cur.execute("""
        INSERT INTO player_groups (player_id, group_id)
        VALUES (%s, %s)
        ON CONFLICT DO NOTHING
    """, (player_id, group_id))
    if cur.rowcount == 1:
        inserted += 1

conn.commit()
print(f"🎯 Fertig: {inserted} neue Player-Groups eingefügt")

# -------- 3) Spieler pro Gruppe abrufen --------
cur.execute("""
    SELECT pg.group_id, pg.player_id
    FROM player_groups pg
    ORDER BY pg.group_id
""")
rows = cur.fetchall()

# Gruppen-zu-Spieler Dictionary
group_to_players = {}
for group_id, player_id in rows:
    group_to_players.setdefault(group_id, []).append(player_id)

# -------- 4) Matches vorbereiten --------
matches_to_insert = []


for group_id, players in group_to_players.items():

    for i in range(len(players)):
        for j in range(i + 1, len(players)):  # nur j > i
            player_id = players[i]
            opponent_id = players[j]
            # Heimspiel
            matches_to_insert.append((player_id, opponent_id, group_id))
            # Auswärtsspiel
            matches_to_insert.append((opponent_id, player_id, group_id))

print(f"🏁 {len(matches_to_insert)} Matches werden vorbereitet...")

# -------- 5) Matches in DB einfügen --------
inserted = 0
for player_id, opponent_id, group_id in matches_to_insert:
    cur.execute("""
        INSERT INTO matches (player_id, opponent_id, group_id, switched_flag)
        VALUES (%s, %s, %s, false)
        ON CONFLICT DO NOTHING
    """, (player_id, opponent_id, group_id))
    if cur.rowcount == 1:
        inserted += 1

conn.commit()
print(f"🎯 Fertig: {inserted} Matches in DB eingefügt.")

cur.close()
conn.close()
